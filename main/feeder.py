import time
from time import sleep
import RPi.GPIO as GPIO
import openpyxl
import shutil
import xlsxwriter
import os

os.chdir('/var/www/html/files')
GPIO.setmode(GPIO.BOARD)
button1 = 16
GPIO.setup(button1,GPIO.IN,pull_up_down=GPIO.PUD_UP)
p=1

while(p==1):
    
    j=13
    name=str(time.strftime('%Y-%m-%d'))+'.xlsx'
    if (os.path.exists(name) == 0):
        workbook = xlsxwriter.Workbook(name)
        workbook.close()

        shutil.copy('default.xlsx',name)

        wb = openpyxl.load_workbook(name)
        sheet = wb.get_sheet_by_name('mnu')
        sheet['C7'] = str(time.strftime('%d %b %Y, %a'))
        wb.save(name)

    while(name==(str(time.strftime('%Y-%m-%d'))+'.xlsx')):
        if GPIO.input(button1)==0 :
            sleep(0.1)
            wb = openpyxl.load_workbook(name)
            sheet1 = wb.get_sheet_by_name('fear')
            sheet = wb.get_sheet_by_name('mnu')
            ref=str(sheet1['A1'].value)
            date=str(sheet1['A2'].value)
            ont=str(sheet1['A3'].value)
            oft=str(sheet1['A4'].value)
            sheet[ref]="NITUK318225"
            sheet[date]=str(time.strftime('%d %b %Y'))
            sheet[ont]=str(time.strftime('%X'))
            sheet[oft]='generator on'
            sheet1['A2']='B'+str(j)
            sheet1['A3']='C'+str(j)
            sheet1['A1']='A'+str(j)
            wb.save(name)
            sleep(0.1)
            print("inside of while 1")
            
            while GPIO.input(button1)==0:
                if(int(str(time.strftime('%X'))[0:2])==23 & int(str(time.strftime('%X'))[3:5])==59 & int(str(time.strftime('%X'))[6:8])>=57):
                    sleep (0.1)
                    break
                print("on")
                sleep(3)
                
            print("outside of while 1")
            sleep(1)
            wb = openpyxl.load_workbook(name)
            sheet1 = wb.get_sheet_by_name('fear')
            sheet = wb.get_sheet_by_name('mnu')
            sheet[oft]=str(time.strftime('%X'))
            sheet1['A4']='D'+str(j)
            sheet['C9']="NITUK318225"
            wb.save(name)
            j=j+1
            sleep(2.1)
        sleep(3)